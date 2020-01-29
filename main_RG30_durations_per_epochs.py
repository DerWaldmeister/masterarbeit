import time
import os
import glob
import numpy as np
import random
import re
import tensorflow as tf
from env import runSimulation, runSimulation_input, activitySequence, activity
from convnet_1d import create1dConvNetNeuralNetworkModel
from convnet_2d import create2dConvNetNeuralNetworkModel
from combined_convnet_2d import createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation
from combined_convnet_1d import createCombined1dConvNetNeuralNetworkModelForFutureResourceUtilisation
import multiprocessing as mp
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from randomize_train_validate_test_indices import randomizeTrainValidateTestIndeces
from datetime import datetime


t_start = time.time()

# user defined parameters
# problem parameters
timeDistribution = "deterministic"    # deterministic, exponential, uniform_1, uniform_2, ...

# file name flag
fileNameLabel = 'main_RG30'

# CPU parameters
numberOfCpuProcessesToGenerateData = 16   # paoloPC has 16 cores
maxTasksPerChildToGenerateData = 4        # 4 is the best for paoloPC

# input state vector  parameters
numberOfActivitiesInStateVector = 6
rescaleFactorTime = 0.1
timeHorizon = 10

# random generation parameters
numberOfSimulationRunsToGenerateData =2000
numberOfSimulationRunsToTestPolicy = 1

# neural network type
neuralNetworkType = "1dimensional convnet" # 1dimensional convnet, 2dimensional convnet, 1dimensional combined convnet, 2dimensional combined convnet
# for 1dimensional convnet and 2dimensional convnet futureResourceUtilisation will not be used
useFutureResourceUtilisation = False
if neuralNetworkType == "1dimensional combined convnet" or neuralNetworkType == "2dimensional combined convnet":
    useFutureResourceUtilisation = True

# train parameter
generateNewTrainTestValidateSets = False
importExistingNeuralNetworkModel = False
neuralNetworkModelAlreadyExists = False
numberOfEpochs = 3000 #walk entire samples
epochsTrainingInterval = 100
# learning rate
learningRate = 0.0001

# test the model on test set
testModelOnTestSet = False

# paths
relativePath = os.path.dirname(__file__)
absolutePathProjects = relativePath + "/database/RG30_Newdata/"

# other parameters
np.set_printoptions(precision=4)    # print precision of numpy variables

# Activity Sequence variables
# initialise variables
numberOfActivities = None
numberOfResources = None
activitySequences = []
decisions_indexActivity = []
decisions_indexActivityPowerset = []

# read all activity sequences from database
absolutePathProjectsGlob = absolutePathProjects + "*.txt"
files = sorted(glob.glob(absolutePathProjectsGlob))

# divide all activity sequences in training and test set
numberOfFiles = len(files)

# call randomizeTrainValidateTestIndeces to generate a new train validate test split or use an already created split
indexFilesTrain, indexFilesValidate, indexFilesTest = randomizeTrainValidateTestIndeces(numberOfFiles, generateNewTrainTestValidateSets , fileNameLabel)

numberOfFilesTrain = len(indexFilesTrain)
numberOfFilesValidate = len(indexFilesValidate)
numberOfFilesTest = len(indexFilesTest)

# organize the read activity sequences in classes
for i in range(numberOfFiles):
    file = files[i]
    #print(file)
    # create a new activitySequence object
    currentActivitySequence = activitySequence()
    with open(file,"r") as f:
        currentActivitySequence.index = i
        currentActivitySequence.fileName = os.path.basename(f.name)
        #print(currentActivitySequence.fileName)
        # allLines = f.read()
        # print(allLines)
        next(f)
        firstLine = f.readline()    # information about numberOfActivities and numberOfResourceTypes
        firstLineDecomposed = re.split(" +", firstLine)
        numberOfActivities = (int(firstLineDecomposed[1])-2)    # the first and last dummy activity do not count
        currentActivitySequence.numberOfActivities = numberOfActivities
        # print("numberOfActivities = " + str(currentActivitySequence.numberOfActivities))
        secondLine = f.readline()   # information about total available resources
        secondLineDecomposed = re.split(" +", secondLine)
        #print(secondLineDecomposed)
        del secondLineDecomposed[0]
        secondLineDecomposed[-1]=secondLineDecomposed[-1].strip('\n')#only string have attributes of strip and delete '\n' at the end
        #print(secondLineDecomposed)
        numberOfResources = 0
        # print(len(secondLineDecomposed))
        #secondLineDecomposed=[int(secondLineDecomposed)]
        secondLineDecomposed = list(map(int,secondLineDecomposed))
        #print(secondLineDecomposed)
        for totalResources in secondLineDecomposed:
            numberOfResources += 1
            #print(numberOfResources)
            currentActivitySequence.totalResources.append(int(totalResources))
            #print(currentActivitySequence.totalResources)
        currentActivitySequence.numberOfResources = numberOfResources
        next(f)
        thirdLine = f.readline()   # information about starting activities
        thirdLineDecomposed = re.split(" +", thirdLine)
        for IdActivity in thirdLineDecomposed[7:-1]:
            currentActivitySequence.indexStartActivities.append(int(IdActivity)-2)
        #print("indexStartingActivities = " + str(currentActivitySequence.indexStartActivities))
        line = f.readline()
        while line:
            #print(line, end="")
            lineDecomposed = re.split(" +", line)
            if int(lineDecomposed[1]) == 0:
                break
            else:
                currentActivity = activity()
                currentActivity.time = int(lineDecomposed[1])
                currentActivity.requiredResources = [ int(lineDecomposed[2]),int(lineDecomposed[3]),int(lineDecomposed[4]),int(lineDecomposed[5])]
                for IdFollowingActivity in lineDecomposed[7:-1]:
                    if int(IdFollowingActivity) != numberOfActivities+2:    #if the following action is not the last dummy activity
                        currentActivity.indexFollowingActivities.append(int(IdFollowingActivity) - 2)
            currentActivitySequence.activities.append(currentActivity)
            line = f.readline()
        #add indexes to list of activities
        for j in range(len(currentActivitySequence.activities)):
            currentActivitySequence.activities[j].index = j
        #add numberOfPreviousActivities
        for Activity in currentActivitySequence.activities:
            for IndexFollowingActivity in Activity.indexFollowingActivities:
                currentActivitySequence.activities[IndexFollowingActivity].numberOfPreviousActivities += 1
    activitySequences.append(currentActivitySequence)

stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources+ timeHorizon * numberOfResources
#print(stateVectorLength)

# compute decisions: each decision corresponds to a start of an activity in the local reference system (more than one decision can be taken at once)
for i in range(0,numberOfActivitiesInStateVector):
    decisions_indexActivity.append(i)


# states, action and futureResourceUtilisationMatrices of training set
states = []
actions = []
futureResourceUtilisationMatrices = []
# states, action and futureResourceUtilisationMatrices of validation set
statesValidationSet = []
actionsValidationSet = []
futureResourceUtilisationMatricesValidationSet = []

# duration records validation and training set
sumTotalDurationRandomValidateRecord = []
sumTotalDurationWithNeuralNetworkModelValidateRecord = []
sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords = []
sumTotalDurationWithCriticalResourceValidateRecord = []
sumTotalDurationWithShortestProcessingValidateRecord = []
sumTotalDurationWithShortestSumDurationValidateRecord = []
sumTotalDurationRandomTrainRecord = []
sumTotalDurationWithNeuralNetworkModelTrainRecord = []
sumTotalDurationWithCriticalResourceTrainRecord = []
sumTotalDurationWithShortestProcessingTrainRecord = []
sumTotalDurationWithShortestSumDurationTrainRecord = []

# test durations
sumTotalDurationRandomTest = 0
sumTotalDurationWithNeuralNetworkModelTest = 0
sumTotalDurationWithCriticalResourceTest = 0
sumTotalDurationWithShortestProcessingTest = 0
sumTotalDurationWithShortestSumDurationTest = 0


#--------------------------------------------------------------RANDOM-----------------------------------------------------------------------------
####  GENERATE TRAINING DATA USING RANDOM DECISIONS (WITHOUT USING pool.map) ####
print('######  RANDOM DECISION ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "generateData"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    # neuralNetworkType needs to be set because states get stored accordingly in states list (vectors vs. matrices)
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)

# assign simulation results to activity sequences and append training data
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesTrain[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesTrain[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesTrain[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesTrain[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesTrain[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean

    # saving validation set states, actions and futureResourceUtilisationMatrices
    for currentStateActionPair in runSimulation_outputs[i].stateActionPairsOfBestRun:
        states.append(currentStateActionPair.state)
        actions.append(currentStateActionPair.action)
        futureResourceUtilisationMatrices.append(currentStateActionPair.futureResourceUtilisationMatrix)


####  CREATE BENCHMARK WITH RANDOM DECISIONS ALSO WITH VALIDATION ACTIVITY SEQUENCES  ####
print('######  RANDOM DECISION ON VALIDATE ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "generateData"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)

# assign simulation results to activity sequences
for i in range(numberOfFilesValidate):
    activitySequences[indexFilesValidate[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesValidate[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesValidate[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesValidate[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesValidate[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesValidate[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean

    # saving validation set states, actions and futureResourceUtilisationMatrices
    for currentStateActionPair in runSimulation_outputs[i].stateActionPairsOfBestRun:
        statesValidationSet.append(currentStateActionPair.state)
        actionsValidationSet.append(currentStateActionPair.action)
        futureResourceUtilisationMatricesValidationSet.append(currentStateActionPair.futureResourceUtilisationMatrix)


#------------------------------------------------------Training neural net-------------------------------------------
####  TRAIN MODEL USING TRAINING DATA  ####
# look for existing model
print("Train neural network model")

# 1dimensional convnet without using futureResoureUtilisationMatrix
if neuralNetworkType == "1dimensional convnet":

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")

        else:
            neuralNetworkModel = create1dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)

    else:
        neuralNetworkModel = create1dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)


    # runId for simulation run
    runId = "1d_config_5_lr" + str(learningRate) + "_epochs" + str(numberOfEpochs)
    # model id for saving the model uniquely
    modelId = datetime.now().strftime('%Y%m%d-%H%M%S')
    epochsCounter = 0
    print("epochsCounter: " + str(epochsCounter))

    # train neuralNet for epochsTrainingInterval number of epochs, calculate ValidateDuration afterwards
    for i in range(int(numberOfEpochs / epochsTrainingInterval)):
        neuralNetworkModel.fit({"input_currentState": states}, {"targets": actions}, n_epoch=epochsTrainingInterval, snapshot_epoch=True, validation_set=(statesValidationSet, actionsValidationSet),
                               show_metric=True, run_id=runId)

        # increment epochsCounter by number of trained epochs
        epochsCounter = epochsCounter + epochsTrainingInterval
        print("epochsCounter: " + str(epochsCounter))

        # save model
        neuralNetworkModel.save('./savedDNN/model' + modelId + '.tfl')

        # duration calculation
        print('###### NEURAL NETWORK MODEL ON VALIDATE ACTIVITY SEQUENCES  ######')
        sumTotalDurationWithNeuralNetworkModelValidate = 0
        for i in range(numberOfFilesValidate):
            currentRunSimulation_input = runSimulation_input()
            currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
            currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
            currentRunSimulation_input.timeDistribution = timeDistribution
            currentRunSimulation_input.purpose = "testPolicy"
            currentRunSimulation_input.randomDecisionProbability = 0
            currentRunSimulation_input.policyType = "neuralNetworkModel"
            currentRunSimulation_input.neuralNetworkType = neuralNetworkType
            currentRunSimulation_input.decisionTool = neuralNetworkModel
            currentRunSimulation_input.numberOfResources = numberOfResources
            currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
            currentRunSimulation_input.stateVectorLength = stateVectorLength
            currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
            currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
            currentRunSimulation_input.numberOfActivities = numberOfActivities
            currentRunSimulation_input.timeHorizon = timeHorizon
            currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

            currentRunSimulation_output = runSimulation(currentRunSimulation_input)

            activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

        # calculates total duration for validation files
        for i in range(numberOfFilesValidate):
            sumTotalDurationWithNeuralNetworkModelValidate += activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy

        # append number of trained epochs and the duration for this number of epochs to Records list
        sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords.append(
            [epochsCounter, sumTotalDurationWithNeuralNetworkModelValidate])

        # load model
        neuralNetworkModel.load('./savedDNN/model' + modelId + '.tfl')


# 2dimensional convnet without using futureResoureUtilisationMatrix
elif neuralNetworkType == "2dimensional convnet":
    # Turn states list into tuples
    states = np.asarray(states)
    # Reshape states
    states = states.reshape([-1, len(states[0]), len(states[0]), 1])

    # Reshape validation states
    statesValidationSet = np.asarray(statesValidationSet)
    statesValidationSet = statesValidationSet.reshape([-1, len(statesValidationSet[0]), len(statesValidationSet[0]), 1])

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            neuralNetworkModel = create2dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
    else:
        neuralNetworkModel = create2dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)

    runId = "2d_config_9_lr" + str(learningRate) + "_epochs" + str(numberOfEpochs)
    # model id for saving the model uniquely
    modelId = datetime.now().strftime('%Y%m%d-%H%M%S')
    epochsCounter = 0
    print("epochsCounter: " + str(epochsCounter))

    # train neuralNet for epochsTrainingInterval number of epochs, calculate ValidateDuration afterwards
    for i in range(int(numberOfEpochs / epochsTrainingInterval)):
        neuralNetworkModel.fit({"input_currentState": states}, {"targets": actions}, n_epoch=epochsTrainingInterval, snapshot_epoch=True, validation_set=(statesValidationSet, actionsValidationSet),
                               show_metric=True, run_id=runId)
        # save model
        neuralNetworkModel.save('./savedDNN/model' + modelId + '.tfl')

        # Increment epochsCounter by number of trained epochs
        epochsCounter = epochsCounter + epochsTrainingInterval
        print("epochsCounter: " + str(epochsCounter))

        # duration calculation for initial training period
        print('###### NEURAL NETWORK MODEL ON VALIDATE ACTIVITY SEQUENCES  ######')
        sumTotalDurationWithNeuralNetworkModelValidate = 0
        for i in range(numberOfFilesValidate):
            currentRunSimulation_input = runSimulation_input()
            currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
            currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
            currentRunSimulation_input.timeDistribution = timeDistribution
            currentRunSimulation_input.purpose = "testPolicy"
            currentRunSimulation_input.randomDecisionProbability = 0
            currentRunSimulation_input.policyType = "neuralNetworkModel"
            currentRunSimulation_input.neuralNetworkType = neuralNetworkType
            currentRunSimulation_input.decisionTool = neuralNetworkModel
            currentRunSimulation_input.numberOfResources = numberOfResources
            currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
            currentRunSimulation_input.stateVectorLength = stateVectorLength
            currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
            currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
            currentRunSimulation_input.numberOfActivities = numberOfActivities
            currentRunSimulation_input.timeHorizon = timeHorizon
            currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

            currentRunSimulation_output = runSimulation(currentRunSimulation_input)

            activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

        # calculates total duration for validation files
        for i in range(numberOfFilesValidate):
            sumTotalDurationWithNeuralNetworkModelValidate += activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy

        # append number of trained epochs and the duration for this number of epochs to Records list
        sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords.append(
            [epochsCounter, sumTotalDurationWithNeuralNetworkModelValidate])

        # load model
        neuralNetworkModel.load('./savedDNN/model' + modelId + '.tfl')


# combination of a 1 dimensional convnet for current state and a 2 dimensional convnet for resourceUtilisationMatrix
elif neuralNetworkType == "1dimensional combined convnet":
    # Turn futureResourceUtilisationMatrices into tuples and reshape afterwards
    futureResourceUtilisationMatrices = np.asarray(futureResourceUtilisationMatrices)
    # Reshape futureResourceUtilisationMatrices, -1: batch_size, height(=rows):len(futureResourceUtilisationMatrices[0]), width(=columns): len(futureResourceUtilisationMatrices[0][0]), channels: 1
    futureResourceUtilisationMatrices = futureResourceUtilisationMatrices.reshape(
        [-1, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]), 1])

    # Turn futureResourceUtilisationMatricesValidationSet into tuples and reshape afterwards
    futureResourceUtilisationMatricesValidationSet = np.asarray(futureResourceUtilisationMatricesValidationSet)
    futureResourceUtilisationMatricesValidationSet = futureResourceUtilisationMatricesValidationSet.reshape(
        [-1, len(futureResourceUtilisationMatricesValidationSet[0]),
         len(futureResourceUtilisationMatricesValidationSet[0][0]), 1])

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            neuralNetworkModel = createCombined1dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),
                                                                                                       len(actions[0]),
                                                                                                       learningRate,
                                                                                                       len(
                                                                                                           futureResourceUtilisationMatrices[
                                                                                                               0]), len(
                    futureResourceUtilisationMatrices[0][0]))

    else:
        neuralNetworkModel = createCombined1dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),
                                                                                                   len(actions[0]),
                                                                                                   learningRate, len(
                futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]))


    runId = "1d_combined_config_4c_lr" + str(learningRate) + "_epochs" + str(numberOfEpochs)
    # model id for saving the model uniquely
    modelId = datetime.now().strftime('%Y%m%d-%H%M%S')
    epochsCounter = 0
    print("epochsCounter: " + str(epochsCounter))

    # train neuralNet for epochsTrainingInterval number of epochs, calculate ValidateDuration afterwards
    for i in range(int(numberOfEpochs / epochsTrainingInterval)):

        neuralNetworkModel.fit({"input_currentState": states,
                                "input_futureResourceUtilisationMatrix": futureResourceUtilisationMatrices},
                               {"targets": actions}, n_epoch=epochsTrainingInterval, validation_set=([statesValidationSet, futureResourceUtilisationMatricesValidationSet], actionsValidationSet), snapshot_epoch=True,
                               show_metric=True, run_id=runId)

        # increment epochsCounter by number of trained epochs
        epochsCounter = epochsCounter + epochsTrainingInterval
        print("epochsCounter: " + str(epochsCounter))

        # save model
        neuralNetworkModel.save('./savedDNN/model' + modelId + '.tfl')

        # duration calculation for initial training period
        print('###### NEURAL NETWORK MODEL ON VALIDATE ACTIVITY SEQUENCES  ######')
        sumTotalDurationWithNeuralNetworkModelValidate = 0
        for i in range(numberOfFilesValidate):
            currentRunSimulation_input = runSimulation_input()
            currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
            currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
            currentRunSimulation_input.timeDistribution = timeDistribution
            currentRunSimulation_input.purpose = "testPolicy"
            currentRunSimulation_input.randomDecisionProbability = 0
            currentRunSimulation_input.policyType = "neuralNetworkModel"
            currentRunSimulation_input.neuralNetworkType = neuralNetworkType
            currentRunSimulation_input.decisionTool = neuralNetworkModel
            currentRunSimulation_input.numberOfResources = numberOfResources
            currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
            currentRunSimulation_input.stateVectorLength = stateVectorLength
            currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
            currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
            currentRunSimulation_input.numberOfActivities = numberOfActivities
            currentRunSimulation_input.timeHorizon = timeHorizon
            currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

            currentRunSimulation_output = runSimulation(currentRunSimulation_input)

            activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

        # calculates total duration for validation files
        for i in range(numberOfFilesValidate):
            sumTotalDurationWithNeuralNetworkModelValidate += activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy

        # append number of trained epochs and the duration for the number of epochs to Records list
        sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords.append(
            [epochsCounter, sumTotalDurationWithNeuralNetworkModelValidate])

        # load model
        neuralNetworkModel.load('./savedDNN/model' + modelId + '.tfl')

# combination of a 2 dimensional convnet for current state and a 2 dimensional convnet for resourceUtilisationMatrix
elif neuralNetworkType == "2dimensional combined convnet":
    # turn states list into tuples
    states = np.asarray(states)
    # reshape states
    states = states.reshape([-1, len(states[0]), len(states[0]), 1])

    # reshape validation states
    statesValidationSet = np.asarray(statesValidationSet)
    statesValidationSet = statesValidationSet.reshape([-1, len(statesValidationSet[0]), len(statesValidationSet[0]), 1])

    # turn futureResourceUtilisationMatrices into tuples
    futureResourceUtilisationMatrices = np.asarray(futureResourceUtilisationMatrices)
    # reshape futureResourceUtilisationMatrices, -1: batch_size, height(=rows):len(futureResourceUtilisationMatrices[0]), width(=columns): len(futureResourceUtilisationMatrices[0][0]), channels: 1
    futureResourceUtilisationMatrices = futureResourceUtilisationMatrices.reshape(
        [-1, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]), 1])

    # turn futureResourceUtilisationMatricesValidationSet into tuples and reshape afterwards
    futureResourceUtilisationMatricesValidationSet = np.asarray(futureResourceUtilisationMatricesValidationSet)
    futureResourceUtilisationMatricesValidationSet = futureResourceUtilisationMatricesValidationSet.reshape(
        [-1, len(futureResourceUtilisationMatricesValidationSet[0]),
         len(futureResourceUtilisationMatricesValidationSet[0][0]), 1])

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            neuralNetworkModel = createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),len(actions[0]),learningRate,len(futureResourceUtilisationMatrices[0]),len(futureResourceUtilisationMatrices[0][0]))
    else:
        neuralNetworkModel = createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),len(actions[0]),learningRate, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]))

    runId = "2d_combined_config_9a2_lr" + str(learningRate) + "_epochs" + str(numberOfEpochs)
    # model id for saving the model uniquely
    modelId = datetime.now().strftime('%Y%m%d-%H%M%S')
    epochsCounter = 0
    print("epochsCounter: " + str(epochsCounter))

    # train neuralNet for epochsTrainingInterval number of epochs, calculate ValidateDuration afterwards
    for i in range(int(numberOfEpochs / epochsTrainingInterval)):

        neuralNetworkModel.fit({"input_currentState": states,
                               "input_futureResourceUtilisationMatrix": futureResourceUtilisationMatrices},
                               {"targets": actions}, n_epoch=epochsTrainingInterval, snapshot_epoch=True, validation_set=([statesValidationSet, futureResourceUtilisationMatricesValidationSet], actionsValidationSet),
                               show_metric=True, run_id=runId)

        # increment epochsCounter by number of trained epochs
        epochsCounter = epochsCounter + epochsTrainingInterval
        print("epochsCounter: " + str(epochsCounter))

        # save model
        neuralNetworkModel.save('./savedDNN/model' + modelId + '.tfl')

        # duration calculation for initial training period
        print('###### NEURAL NETWORK MODEL ON VALIDATE ACTIVITY SEQUENCES  ######')
        sumTotalDurationWithNeuralNetworkModelValidate = 0
        for i in range(numberOfFilesValidate):
            currentRunSimulation_input = runSimulation_input()
            currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
            currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
            currentRunSimulation_input.timeDistribution = timeDistribution
            currentRunSimulation_input.purpose = "testPolicy"
            currentRunSimulation_input.randomDecisionProbability = 0
            currentRunSimulation_input.policyType = "neuralNetworkModel"
            currentRunSimulation_input.neuralNetworkType = neuralNetworkType
            currentRunSimulation_input.decisionTool = neuralNetworkModel
            currentRunSimulation_input.numberOfResources = numberOfResources
            currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
            currentRunSimulation_input.stateVectorLength = stateVectorLength
            currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
            currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
            currentRunSimulation_input.numberOfActivities = numberOfActivities
            currentRunSimulation_input.timeHorizon = timeHorizon
            currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

            currentRunSimulation_output = runSimulation(currentRunSimulation_input)

            activitySequences[indexFilesValidate[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

        # calculates total duration for validation files
        for i in range(numberOfFilesValidate):
            sumTotalDurationWithNeuralNetworkModelValidate += activitySequences[
                indexFilesValidate[i]].totalDurationWithPolicy

        # append number of trained epochs and the duration for this number of epochs to Records list
        sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords.append(
            [epochsCounter, sumTotalDurationWithNeuralNetworkModelValidate])

        # load model
        neuralNetworkModel.load('./savedDNN/model' + modelId + '.tfl')


#### RANDOM DECISION ON TEST ACTIVITY SEQUENCES ####
if testModelOnTestSet:
    # RANDOM DECISION
    print('######  RANDOM DECISION ON TEST ACTIVITY SEQUENCES  ######')
    runSimulation_inputs = []
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 1
        currentRunSimulation_input.policyType = None
        currentRunSimulation_input.neuralNetworkType = neuralNetworkType
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities
        currentRunSimulation_input.timeHorizon = timeHorizon
        currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[indexFilesTest[i]].totalDurationMean = currentRunSimulation_output.totalDurationMean
        print("activitySequences[indexFilesTest[i]].totalDurationWithMean: " + str(
            activitySequences[indexFilesTest[i]].totalDurationMean))


    # calculates total duration for test files
    for i in range(numberOfFilesTest):
        sumTotalDurationRandomTest += activitySequences[
            indexFilesTest[i]].totalDurationMean


#-----------------------------------------------------------------NeuralNet------------------------------------------------------------------------------
####  TEST NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTrain[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean


####  TEST NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ####
if testModelOnTestSet:
    print('###### NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ######')
    sumTotalDurationWithNeuralNetworkModelValidate = 0
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "neuralNetworkModel"
        currentRunSimulation_input.neuralNetworkType = neuralNetworkType
        currentRunSimulation_input.decisionTool = neuralNetworkModel
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities
        currentRunSimulation_input.timeHorizon = timeHorizon
        currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[indexFilesTest[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

    # calculates total duration for test files
    for i in range(numberOfFilesTest):
        sumTotalDurationWithNeuralNetworkModelTest += activitySequences[
            indexFilesTest[i]].totalDurationWithPolicy



# code for priority rules from: https://github.com/leiiiiii/RCPSP/blob/master/Env.py

#---------------------------------------------------------Critical Resource----------------------------------------------------------------------------
    ####  TEST CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource = runSimulation_outputs[i].totalDurationMean


####  TEST CRITICAL RESOURCE METHOD ON VALIDATE ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithCriticalResource = currentRunSimulation_output.totalDurationMean

if testModelOnTestSet:
    print('###### CRITICAL RESOURCE METHOD ON TEST ACTIVITY SEQUENCES  ######')
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "most critical resource"
        currentRunSimulation_input.neuralNetworkType = None
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities
        currentRunSimulation_input.timeHorizon = timeHorizon
        currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[
            indexFilesTest[i]].totalDurationWithCriticalResource = currentRunSimulation_output.totalDurationMean


# ---------------------------------------------------------Shortest Processing Time----------------------------------------------------------------------------
####  TEST SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST PROCESSING TIME METHOD ON VALIDATE ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithShortestProcessingTime = currentRunSimulation_output.totalDurationMean

if testModelOnTestSet:
    print('###### SHORTEST PROCESSING TIME METHOD ON TEST ACTIVITY SEQUENCES  ######')
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "shortest processing time"
        currentRunSimulation_input.neuralNetworkType = None
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities
        currentRunSimulation_input.timeHorizon = timeHorizon
        currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[
            indexFilesTest[
                i]].totalDurationWithShortestProcessingTime = currentRunSimulation_output.totalDurationMean

 # ---------------------------------------------------------shortest sumDuration including successor----------------------------------------------------------------------------
####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR TIME METHOD ON VALIDATE ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon
    currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithShortestSumDuration = currentRunSimulation_output.totalDurationMean

if testModelOnTestSet:
    print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TEST ACTIVITY SEQUENCES  ######')
    for i in range(numberOfFilesTest):
        currentRunSimulation_input = runSimulation_input()
        currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
        currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
        currentRunSimulation_input.timeDistribution = timeDistribution
        currentRunSimulation_input.purpose = "testPolicy"
        currentRunSimulation_input.randomDecisionProbability = 0
        currentRunSimulation_input.policyType = "shortest sumDuration including successor"
        currentRunSimulation_input.neuralNetworkType = None
        currentRunSimulation_input.decisionTool = None
        currentRunSimulation_input.numberOfResources = numberOfResources
        currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
        currentRunSimulation_input.stateVectorLength = stateVectorLength
        currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
        currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
        currentRunSimulation_input.numberOfActivities = numberOfActivities
        currentRunSimulation_input.timeHorizon = timeHorizon
        currentRunSimulation_input.useFutureResourceUtilisation = useFutureResourceUtilisation

        currentRunSimulation_output = runSimulation(currentRunSimulation_input)

        activitySequences[
            indexFilesTest[i]].totalDurationWithShortestSumDuration = currentRunSimulation_output.totalDurationMean


#------------------------------------------------------EVALUATION-----------------------------------------------------------------------------
####  EVALUATION OF RESULTS OF TRAIN ACTIVITY SEQUENCES  ####
sumTotalDurationRandomTrain = 0
sumTotalDurationWithCriticalResourceTrain = 0
sumTotalDurationWithShortestProcessingTrain = 0
sumTotalDurationWithNeuralNetworkModelTrain = 0
sumTotalDurationWithShortestSumDurationTrain = 0

for i in range(numberOfFilesTrain):
    sumTotalDurationRandomTrain += activitySequences[indexFilesTrain[i]].totalDurationMean
    sumTotalDurationRandomTrain = round(sumTotalDurationRandomTrain,4)
    sumTotalDurationWithNeuralNetworkModelTrain += activitySequences[indexFilesTrain[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceTrain += activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration

sumTotalDurationRandomTrainRecord.append(sumTotalDurationRandomTrain)
sumTotalDurationWithNeuralNetworkModelTrainRecord.append(sumTotalDurationWithNeuralNetworkModelTrain)
sumTotalDurationWithCriticalResourceTrainRecord.append(sumTotalDurationWithCriticalResourceTrain)
sumTotalDurationWithShortestProcessingTrainRecord.append(sumTotalDurationWithShortestProcessingTrain)
sumTotalDurationWithShortestSumDurationTrainRecord.append(sumTotalDurationWithShortestSumDurationTrain)

####  EVALUATION OF NN RESULTS OF VALIDATE ACTIVITY SEQUENCES  ####
sumTotalDurationRandomValidate = 0
sumTotalDurationWithCriticalResourceValidate = 0
sumTotalDurationWithShortestProcessingValidate = 0
sumTotalDurationWithShortestSumDurationValidate = 0

for i in range(numberOfFilesValidate):
    sumTotalDurationRandomValidate += activitySequences[indexFilesValidate[i]].totalDurationMean
    sumTotalDurationRandomValidate = round(sumTotalDurationRandomValidate,4)
    sumTotalDurationWithCriticalResourceValidate += activitySequences[indexFilesValidate[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingValidate += activitySequences[indexFilesValidate[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationValidate += activitySequences[indexFilesValidate[i]].totalDurationWithShortestSumDuration


sumTotalDurationRandomValidateRecord.append(sumTotalDurationRandomValidate)
sumTotalDurationWithCriticalResourceValidateRecord.append(sumTotalDurationWithCriticalResourceValidate)
sumTotalDurationWithShortestProcessingValidateRecord.append(sumTotalDurationWithShortestProcessingValidate)
sumTotalDurationWithShortestSumDurationValidateRecord.append(sumTotalDurationWithShortestSumDurationValidate)


####  EVALUATION OF DURATIONS OF TEST ACTIVITY SEQUENCES  ####
if testModelOnTestSet:
    sumTotalDurationRandomTest = 0
    sumTotalDurationWithNeuralNetworkModelTest = 0
    sumTotalDurationWithCriticalResourceTest = 0
    sumTotalDurationWithShortestProcessingTest = 0
    sumTotalDurationWithShortestSumDurationTest = 0

    for i in range(numberOfFilesTest):
        sumTotalDurationRandomTest += activitySequences[indexFilesTest[i]].totalDurationMean
        sumTotalDurationRandomTest = round(sumTotalDurationRandomTest,4)
        sumTotalDurationWithNeuralNetworkModelTest += activitySequences[indexFilesTest[i]].totalDurationWithPolicy
        sumTotalDurationWithCriticalResourceTest += activitySequences[indexFilesTest[i]].totalDurationWithCriticalResource
        sumTotalDurationWithShortestProcessingTest += activitySequences[indexFilesTest[i]].totalDurationWithShortestProcessingTime
        sumTotalDurationWithShortestSumDurationTest += activitySequences[indexFilesTest[i]].totalDurationWithShortestSumDuration




print("neuralNetworkType: " + neuralNetworkType)
print("sumTotalDurationRandomTrain = " + str(sumTotalDurationRandomTrain))
print("sumTotalDurationWithNeuralNetworkModelTrain = " + str(sumTotalDurationWithNeuralNetworkModelTrain))
print("sumTotalDurationWithCriticalResourceTrain = " + str(sumTotalDurationWithCriticalResourceTrain))
print("sumTotalDurationWithShortestProcessingTrain = " + str(sumTotalDurationWithShortestProcessingTrain))
print("sumTotalDurationWithShortestSumDurationTrain = " + str(sumTotalDurationWithShortestSumDurationTrain))
print("sumTotalDurationRandomValidate = " + str(sumTotalDurationRandomValidate))
print("sumTotalDurationWithCriticalResourceValidate = " + str(sumTotalDurationWithCriticalResourceValidate))
print("sumTotalDurationWithShortestProcessingValidate = " + str(sumTotalDurationWithShortestProcessingValidate))
print("sumTotalDurationWithShortestSumDurationValidate = " + str(sumTotalDurationWithShortestSumDurationValidate))
# results on test set activities are written into excel sheet

# compute computation time
t_end = time.time()
t_computation = t_end - t_start
print("t_computation = " + str(t_computation))


# write ouput to excel
wb = Workbook()
ws = wb.create_sheet('Durations_rg30',0)

alignCenter = Alignment(horizontal='center')

ws['A1'] = runId
ws['C1'] = 'LR:'
ws['D1'].value = learningRate
ws['E1'] = 'epochs:'
ws['F1'].value = numberOfEpochs
ws['A2'] = 'Durations'
ws['B2'] = 'Computation time'
ws['C2'].value = t_computation
ws.merge_cells('A3:E3')
ws.merge_cells('F3:J3')
ws['A3'] = 'durations on train set'
ws['A3'].alignment = alignCenter
ws['F3'] = 'durations on validate set'
ws['F3'].alignment = alignCenter

ws['A4'] = 'Random'
ws['B4'] = 'NeuralNetworkModel'
ws['C4'] = 'CriticalResource'
ws['D4'] = 'ShortestProcessing'
ws['E4'] = 'ShortestSumDuration'
ws['F4'] = 'Random'
ws['G4'] = 'NeuralNetworkModel'
ws['H4'] = 'CriticalResource'
ws['I4'] = 'ShortestProcessing'
ws['J4'] = 'ShortestSumDuration'

ws['A5'].value = sumTotalDurationRandomTrain
ws['B5'].value = sumTotalDurationWithNeuralNetworkModelTrain
ws['C5'].value = sumTotalDurationWithCriticalResourceTrain
ws['D5'].value = sumTotalDurationWithShortestProcessingTrain
ws['E5'].value = sumTotalDurationWithShortestSumDurationTrain
ws['F5'].value = sumTotalDurationRandomValidate
ws['H5'].value = sumTotalDurationWithCriticalResourceValidate
ws['I5'].value = sumTotalDurationWithShortestProcessingValidate
ws['J5'].value = sumTotalDurationWithShortestSumDurationValidate

ws['A7'] = 'Epochs:'
ws['A8'] = 'DurationNNValidate'
ws['A9'] = 'Percent. improv. NN vs. Random'

for columnPos in range(len(sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords)):
    # epochs
    ws.cell(column=2+columnPos, row=7, value=str(sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords[columnPos][0]))
    # durations
    ws.cell(column=2+columnPos, row=8, value=str(sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords[columnPos][1]))

    # calculate percentage improvement of NNValidateDuration vs RandomValidateDuration
    percentageImprovement = round(100*(1 - (sumTotalDurationsPerEpochsWithNeuralNetworkModelValidateRecords[columnPos][1]/sumTotalDurationRandomValidate)), 2)
    # write it in excel sheet
    ws.cell(column=2+columnPos, row=9,
            value=str(percentageImprovement))

# durations on test set
if testModelOnTestSet:
    ws['A11'] = 'durations on test set'
    ws['A12'] = 'Random'
    ws['B12'] = 'NeuralNetworkModel'
    ws['C12'] = 'CriticalResource'
    ws['D12'] = 'ShortestProcessing'
    ws['E12'] = 'ShortestSumDuration'

    ws['A13'].value = sumTotalDurationRandomTest
    ws['B13'].value = sumTotalDurationWithNeuralNetworkModelTest
    ws['C13'].value = sumTotalDurationWithCriticalResourceTest
    ws['D13'].value = sumTotalDurationWithShortestProcessingTest
    ws['E13'].value = sumTotalDurationWithShortestSumDurationTest


ws.column_dimensions['A'].width = 10.0
ws.column_dimensions['B'].width = 18.0
ws.column_dimensions['C'].width = 14.0
ws.column_dimensions['D'].width = 18.0
ws.column_dimensions['E'].width = 19.0
ws.column_dimensions['F'].width = 10.0
ws.column_dimensions['G'].width = 18.0
ws.column_dimensions['H'].width = 14.0
ws.column_dimensions['I'].width = 18.0
ws.column_dimensions['J'].width = 19.0

ws['A4'].alignment = alignCenter
ws['B4'].alignment = alignCenter
ws['C4'].alignment = alignCenter
ws['D4'].alignment = alignCenter
ws['E4'].alignment = alignCenter
ws['F4'].alignment = alignCenter
ws['G4'].alignment = alignCenter
ws['H4'].alignment = alignCenter
ws['I4'].alignment = alignCenter
ws['J4'].alignment = alignCenter

ws['A5'].alignment = alignCenter
ws['B5'].alignment = alignCenter
ws['C5'].alignment = alignCenter
ws['D5'].alignment = alignCenter
ws['E5'].alignment = alignCenter
ws['F5'].alignment = alignCenter
ws['G5'].alignment = alignCenter
ws['H5'].alignment = alignCenter
ws['I5'].alignment = alignCenter
ws['J5'].alignment = alignCenter

logdir = '/durations/' + neuralNetworkType +'/' + datetime.now().strftime('%Y%m%d-%H%M%S') + '_durations_rg30.xlsx'

wb.save(relativePath + logdir)
