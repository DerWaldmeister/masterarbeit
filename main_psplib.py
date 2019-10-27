import time
import os
import glob
import numpy as np
import tensorflow as tf
import random
import re
from randomize_train_validate_test_indices import randomizeTrainValidateTestIndeces
from env import runSimulation, runSimulation_input, activitySequence, activity
from convnet_1d import create1dConvNetNeuralNetworkModel
from convnet_2d import create2dConvNetNeuralNetworkModel
from convnet_2d_futureResourceUtilisation import createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation

import multiprocessing as mp
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side

t_start = time.time()

# user defined parameters
# problem parameters
timeDistribution = "deterministic"    # deterministic, exponential, uniform_1, uniform_2, ...#

# file name flag
fileNameLabel = 'main_psplib'

# CPU parameters
numberOfCpuProcessesToGenerateData = 16   # paoloPC has 16 cores
maxTasksPerChildToGenerateData = 4        # 4 is the best for paoloPC

# input state vector  parameters
numberOfActivitiesInStateVector = 6
rescaleFactorTime = 0.1
timeHorizon = 10

# random generation parameters
numberOfSimulationRunsToGenerateData = 2000
numberOfSimulationRunsToTestPolicy = 1
numberOfMainRun = 1

# neural network type
neuralNetworkType = "2dimensional combined convnet"   # 1dimensional convnet, 2dimensional convnet, 1dimensional combined convnet, 2dimensional combined convnet
# for 1dimensional convnet and 2dimesnional convnet futureResourceUtilisation wont be used
if neuralNetworkType == "1dimensional convnet" or neuralNetworkType == "2dimensional convnet":
    useFutureResourceUtilisation = False

# train parameters
generateNewTrainTestValidateSets = False
importExistingNeuralNetworkModel = False
neuralNetworkModelAlreadyExists = False
numberOfEpochs = 50 #walk entire samples
learningRate = 0.005

# paths
relativePath = os.path.dirname(__file__)
absolutePathProjects = relativePath + "/database/psplib_J30/"

# other parameters
np.set_printoptions(precision=4)    # print precision of numpy variables

# initialise variables
numberOfActivities = None
numberOfResources = None
activitySequences = []
decisions_indexActivity = []
decisions_indexActivityPowerset = []
states = []
actions = []
futureResourceUtilisationMatrices = []
#actionsPossibilities = []
sumTotalDurationRandomTestRecord = []
sumTotalDurationWithNeuralNetworkModelTestRecord = []
sumTotalDurationWithCriticalResourceTestRecord = []
sumTotalDurationWithShortestProcessingTestRecord = []
sumTotalDurationWithShortestSumDurationTestRecord = []
sumTotalDurationRandomTrainRecord = []
sumTotalDurationWithNeuralNetworkModelTrainRecord = []
sumTotalDurationWithCriticalResourceTrainRecord = []
sumTotalDurationWithShortestProcessingTrainRecord = []
sumTotalDurationWithShortestSumDurationTrainRecord = []

# read all activity sequences from database
absolutePathProjectsGlob = absolutePathProjects + "*.txt"
files = sorted(glob.glob(absolutePathProjectsGlob))

# divide all activity sequences in training and test set
numberOfFiles = len(files)  # 480
#print("numberOfFiles:" + str(numberOfFiles))

indexFilesTrain, indexFilesValidate, indexFilesTest = randomizeTrainValidateTestIndeces(numberOfFiles, generateNewTrainTestValidateSets , fileNameLabel)

numberOfFilesTrain = len(indexFilesTrain)
numberOfFilesValidate = len(indexFilesValidate)
numberOfFilesTest = len(indexFilesTest)

'''
numberOfFilesTest = round(numberOfFiles * percentageOfFilesTest)
numberOfFilesTrain = numberOfFiles - numberOfFilesTest
indexFiles = list(range(0, numberOfFiles))
indexFilesTrain = []
indexFilesTest = []

# choose the first element of every set to test
for i in range(numberOfFilesTest):
    # randomIndex = random.randrange(0, len(indexFiles))
    randomIndex = i*9
    indexFilesTest.append(indexFiles[randomIndex])
    del indexFiles[randomIndex]#delete
indexFilesTrain = indexFiles
'''

# organize the read activity sequences in classes
for i in range(numberOfFiles):
    file = files[i]
    # create a new activitySequence object
    currentActivitySequence = activitySequence()
    with open(file, "r") as f:
        currentActivitySequence.index = i
        currentActivitySequence.fileName = os.path.basename(f.name)
        firstLine = f.readline()
        firstLineDecomposed = re.split(" +", firstLine)
        numberOfActivities = (int(firstLineDecomposed[0]) - 2)
        currentActivitySequence.numberOfActivities = numberOfActivities
        secondLine = f.readline()
        secondLineDecomposed = re.split(" +", secondLine)
        numberOfResources = 0
        for totalResources in secondLineDecomposed[0:-1]:
            numberOfResources += 1
            currentActivitySequence.totalResources.append(int(totalResources))
        currentActivitySequence.numberOfResources = numberOfResources
        thirdLine = f.readline()
        thirdLineDecomposed = re.split(" +", thirdLine)
        for IdActivity in thirdLineDecomposed[6:-1]:
            currentActivitySequence.indexStartActivities.append(int(IdActivity) - 2)
        line = f.readline()
        while line:
            lineDecomposed = re.split(" +", line)
            if int(lineDecomposed[0]) == 0:
                break
            else:
                currentActivity = activity()
                currentActivity.time = int(lineDecomposed[0])
                currentActivity.requiredResources = [int(lineDecomposed[1]), int(lineDecomposed[2]),int(lineDecomposed[3]), int(lineDecomposed[4])]
                for IdFollowingActivity in lineDecomposed[6:-1]:
                    if int(IdFollowingActivity) != numberOfActivities + 2:  # if the following action is not the last dummy activity
                        currentActivity.indexFollowingActivities.append(int(IdFollowingActivity) - 2)
            currentActivitySequence.activities.append(currentActivity)
            line = f.readline()
        # add indexes to list of activities
        for j in range(len(currentActivitySequence.activities)):
            currentActivitySequence.activities[j].index = j
        # add numberOfPreviousActivities
        for Activity in currentActivitySequence.activities:
            for IndexFollowingActivity in Activity.indexFollowingActivities:
                currentActivitySequence.activities[IndexFollowingActivity].numberOfPreviousActivities += 1
    activitySequences.append(currentActivitySequence)

# So far no time horizon in the input vector (neither in psplib nor in RG30)
#stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources + timeHorizon * numberOfResources
stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources

# compute decisions: each decision corresponds to a start of an activity in the local reference system (more than one decision can be taken at once)
for i in range(0, numberOfActivitiesInStateVector):
    decisions_indexActivity.append(i)



#--------------------------------------------------------------RANDOM-----------------------------------------------------------------------------
####  GENERATE TRAINING DATA USING RANDOM DECISIONS (WITHOUT USING pool.map) ####
print("######  RANDOM DECISION ON TRAIN ACTIVITY SEQUENCES  ######")
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    #print("numberOfSimulationRunsToGenerateData: " + str(numberOfSimulationRunsToGenerateData))
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "generateData"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    # neuralNetworkType needs to be set because states get stored accordingly in states list (vectors vs. matrices)
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences and append training data



for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesTrain[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesTrain[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesTrain[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesTrain[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesTrain[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean

    for currentStateActionPair in runSimulation_outputs[i].stateActionPairsOfBestRun:
        states.append(currentStateActionPair.state)
        actions.append(currentStateActionPair.action)
        futureResourceUtilisationMatrices.append(currentStateActionPair.futureResourceUtilisationMatrix)


####  TRAIN MODEL USING TRAINING DATA  ####
# look for existing model
print("Train neural network model")

# 1dimensional convnet without using futureResoureUtilisationMatrix
if neuralNetworkType == "1dimensional convnet":
    if importExistingNeuralNetworkModel:
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")

        else:
            neuralNetworkModel = create1dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
            #neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actionsPossibilities[0]), learningRate)
    else:
        neuralNetworkModel = create1dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
        #neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actionsPossibilities[0]), learningRate)

    neuralNetworkModel.fit({"input": states}, {"targets": actions}, n_epoch=numberOfEpochs, snapshot_epoch=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")

# 2dimensional convnet without using futureResoureUtilisationMatrix
elif neuralNetworkType == "2dimensional convnet":
    # Turn states list into tuples
    states = np.asarray(states)
    #print("states: " + str(states))
    #print("states[0]: " + str(states[0]))
    # Reshape states
    states = states.reshape([-1, len(states[0]), len(states[0]), 1])
    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            neuralNetworkModel = create2dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
    else:
        neuralNetworkModel = create2dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)

    neuralNetworkModel.fit({"input": states}, {"targets": actions}, n_epoch=numberOfEpochs, snapshot_epoch=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")


elif neuralNetworkType == "1dimensional combined convnet":
    placeholder = True

elif neuralNetworkType == "2dimensional combined convnet":
    # Turn states list into tuples
    states = np.asarray(states)
    #print("states: " + str(states))
    #print("states[0]: " + str(states[0]))
    # Reshape states
    states = states.reshape([-1, len(states[0]), len(states[0]), 1])

    # NEW:
    # Turn futureResourceUtilisationMatrices into tuples
    futureResourceUtilisationMatrices = np.asarray(futureResourceUtilisationMatrices)
    # Reshape futureResourceUtilisationMatrices, -1: batch_size, height(=rows):len(futureResourceUtilisationMatrices[0]), width(=columns): len(futureResourceUtilisationMatrices[0][0]), channels: 1
    futureResourceUtilisationMatrices = futureResourceUtilisationMatrices.reshape([-1, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]), 1])

    #print("len(futureResourceUtilisationMatrices[0]): " + str(len(futureResourceUtilisationMatrices[0])))
    #print("len(futureResourceUtilisationMatrices[0][0]): " + str(len(futureResourceUtilisationMatrices[0][0])))
    #print("futureResourceUtilisationMatrices: " + str(futureResourceUtilisationMatrices))

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            # NEW:
            neuralNetworkModel = createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),len(actions[0]),learningRate,len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]))
    else:
        # NEW:
        neuralNetworkModel = createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),len(actions[0]),learningRate, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]))

    # NEW:
    neuralNetworkModel.fit({"input_currentState": states,
                           "input_futureResourceUtilisationMatrix": futureResourceUtilisationMatrices},
                           {"targets": actions}, n_epoch=numberOfEpochs, snapshot_epoch=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")
else:
    print("No neural network")

####  CREATE BENCHMARK WITH RANDOM DECISIONS ALSO WITH TEST ACTIVITY SEQUENCES  ####
print('######  RANDOM DECISION ON TEST ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences

for i in range(numberOfFilesTest):
    activitySequences[indexFilesTest[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesTest[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesTest[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesTest[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesTest[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesTest[i]].trivialDecisionPercentageMean = runSimulation_outputs[
        i].trivialDecisionPercentageMean


# -----------------------------------------------------------------NN------------------------------------------------------------------------------
####  TEST NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTrain[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

####  TEST NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean


# Xialoei: critical resource method, shortest processing time, shortest sumDuration

#---------------------------------------------------------Critical Resource----------------------------------------------------------------------------
    ####  TEST CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource = runSimulation_outputs[i].totalDurationMean


####  TEST CRITICAL RESOURCE METHOD ON TEST ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithCriticalResource = currentRunSimulation_output.totalDurationMean

# ---------------------------------------------------------Shortest Processing Time----------------------------------------------------------------------------
####  TEST SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST PROCESSING TIME METHOD ON TEST ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithShortestProcessingTime = currentRunSimulation_output.totalDurationMean

 # ---------------------------------------------------------shortest sumDuration including successor----------------------------------------------------------------------------
####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR TIME METHOD ON TEST ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TEST ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTest):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTest[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTest[i]].totalDurationWithShortestSumDuration = currentRunSimulation_output.totalDurationMean

#------------------------------------------------------EVALUATION-----------------------------------------------------------------------------
####  EVALUATION OF RESULTS OF TRAIN ACTIVITY SEQUENCES  ####
sumTotalDurationRandomTrain = 0
sumTotalDurationWithCriticalResourceTrain = 0
sumTotalDurationWithShortestProcessingTrain = 0
sumTotalDurationWithNeuralNetworkModelTrain = 0
sumTotalDurationWithShortestSumDurationTrain = 0

for i in range(numberOfFilesTrain):
    sumTotalDurationRandomTrain += activitySequences[indexFilesTrain[i]].totalDurationMean
    sumTotalDurationRandomTrain = round(sumTotalDurationRandomTrain,4)
    sumTotalDurationWithNeuralNetworkModelTrain += activitySequences[indexFilesTrain[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceTrain += activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration

sumTotalDurationRandomTrainRecord.append(sumTotalDurationRandomTrain)
sumTotalDurationWithNeuralNetworkModelTrainRecord.append(sumTotalDurationWithNeuralNetworkModelTrain)
sumTotalDurationWithCriticalResourceTrainRecord.append(sumTotalDurationWithCriticalResourceTrain)
sumTotalDurationWithShortestProcessingTrainRecord.append(sumTotalDurationWithShortestProcessingTrain)
sumTotalDurationWithShortestSumDurationTrainRecord.append(sumTotalDurationWithShortestSumDurationTrain)

####  EVALUATION OF NN RESULTS OF TEST ACTIVITY SEQUENCES  ####
sumTotalDurationRandomTest = 0
sumTotalDurationWithNeuralNetworkModelTest = 0
sumTotalDurationWithCriticalResourceTest = 0
sumTotalDurationWithShortestProcessingTest = 0
sumTotalDurationWithShortestSumDurationTest = 0

for i in range(numberOfFilesTest):
    sumTotalDurationRandomTest += activitySequences[indexFilesTest[i]].totalDurationMean
    sumTotalDurationRandomTest = round(sumTotalDurationRandomTest,4)
    sumTotalDurationWithNeuralNetworkModelTest += activitySequences[indexFilesTest[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceTest += activitySequences[indexFilesTest[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingTest += activitySequences[indexFilesTest[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationTest += activitySequences[indexFilesTest[i]].totalDurationWithShortestSumDuration


sumTotalDurationRandomTestRecord.append(sumTotalDurationRandomTest)
sumTotalDurationWithNeuralNetworkModelTestRecord.append(sumTotalDurationWithNeuralNetworkModelTest)
sumTotalDurationWithCriticalResourceTestRecord.append(sumTotalDurationWithCriticalResourceTest)
sumTotalDurationWithShortestProcessingTestRecord.append(sumTotalDurationWithShortestProcessingTest)
sumTotalDurationWithShortestSumDurationTestRecord.append(sumTotalDurationWithShortestSumDurationTest)


print("neuralNetworkType: " + neuralNetworkType)
print("sumTotalDurationRandomTrain = " + str(sumTotalDurationRandomTrain))
print("sumTotalDurationWithNeuralNetworkModelTrain = " + str(sumTotalDurationWithNeuralNetworkModelTrain))
print("sumTotalDurationWithCriticalResourceTrain = " + str(sumTotalDurationWithCriticalResourceTrain))
print("sumTotalDurationWithShortestProcessingTrain = " + str(sumTotalDurationWithShortestProcessingTrain))
print("sumTotalDurationWithShortestSumDurationTrain = " + str(sumTotalDurationWithShortestSumDurationTrain))
print("sumTotalDurationRandomTest = " + str(sumTotalDurationRandomTest))
print("sumTotalDurationWithNeuralNetworkModelTest = " + str(sumTotalDurationWithNeuralNetworkModelTest))
print("sumTotalDurationWithCriticalResourceTest = " + str(sumTotalDurationWithCriticalResourceTest))
print("sumTotalDurationWithShortestProcessingTest = " + str(sumTotalDurationWithShortestProcessingTest))
print("sumTotalDurationWithShortestSumDurationTest = " + str(sumTotalDurationWithShortestSumDurationTest))


# compute computation time
t_end = time.time()
t_computation = t_end - t_start
print("t_computation = " + str(t_computation))



#write ouput to excel
wb = Workbook()
ws = wb.create_sheet('Durations_psplib',0)

alignCenter = Alignment(horizontal='center')

ws['A1'] = 'Durations'
ws['B1'] = 'Computation time'
ws['C1'].value = t_computation
ws.merge_cells('A2:E2')
ws.merge_cells('F2:J2')
ws['A2'] = 'durations on train set'
ws['A2'].alignment = alignCenter
ws['F2'] = 'durations on test set'
ws['F2'].alignment = alignCenter

ws['A3'] = 'Random'
ws['B3'] = 'NeuralNetworkModel'
ws['C3'] = 'CriticalResource'
ws['D3'] = 'ShortestProcessing'
ws['E3'] = 'ShortestSumDuration'
ws['F3'] = 'Random'
ws['G3'] = 'NeuralNetworkModel'
ws['H3'] = 'CriticalResource'
ws['I3'] = 'ShortestProcessing'
ws['J3'] = 'ShortestSumDuration'

ws['A4'].value = sumTotalDurationRandomTrain
ws['B4'].value = sumTotalDurationWithNeuralNetworkModelTrain
ws['C4'].value = sumTotalDurationWithCriticalResourceTrain
ws['D4'].value = sumTotalDurationWithShortestProcessingTrain
ws['E4'].value = sumTotalDurationWithShortestSumDurationTrain
ws['F4'].value = sumTotalDurationRandomTest
ws['G4'].value = sumTotalDurationWithNeuralNetworkModelTest
ws['H4'].value = sumTotalDurationWithCriticalResourceTest
ws['I4'].value = sumTotalDurationWithShortestProcessingTest
ws['J4'].value = sumTotalDurationWithShortestSumDurationTest




ws.column_dimensions['A'].width = 10.0
ws.column_dimensions['B'].width = 18.0
ws.column_dimensions['C'].width = 14.0
ws.column_dimensions['D'].width = 18.0
ws.column_dimensions['E'].width = 19.0
ws.column_dimensions['F'].width = 10.0
ws.column_dimensions['G'].width = 18.0
ws.column_dimensions['H'].width = 14.0
ws.column_dimensions['I'].width = 18.0
ws.column_dimensions['J'].width = 19.0

ws['A3'].alignment = alignCenter
ws['B3'].alignment = alignCenter
ws['C3'].alignment = alignCenter
ws['D3'].alignment = alignCenter
ws['E3'].alignment = alignCenter
ws['F3'].alignment = alignCenter
ws['G3'].alignment = alignCenter
ws['H3'].alignment = alignCenter
ws['I3'].alignment = alignCenter
ws['J3'].alignment = alignCenter

ws['A4'].alignment = alignCenter
ws['B4'].alignment = alignCenter
ws['C4'].alignment = alignCenter
ws['D4'].alignment = alignCenter
ws['E4'].alignment = alignCenter
ws['F4'].alignment = alignCenter
ws['G4'].alignment = alignCenter
ws['H4'].alignment = alignCenter
ws['I4'].alignment = alignCenter
ws['J4'].alignment = alignCenter


wb.save(relativePath + "/durations_psplib.xlsx")
