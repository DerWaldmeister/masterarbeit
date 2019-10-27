import time
import os
import glob
import numpy as np
import random
import re
from env import runSimulation, runSimulation_input, activitySequence, activity
from convnet_1d import create1dConvNetNeuralNetworkModel
from convnet_2d import create2dConvNetNeuralNetworkModel
from convnet_2d_futureResourceUtilisation import createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation
import multiprocessing as mp
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side
from randomize_train_validate_test_indices import randomizeTrainValidateTestIndeces

t_start = time.time()

# user defined parameters
# problem parameters
timeDistribution = "deterministic"    # deterministic, exponential, uniform_1, uniform_2, ...

# file name flag
fileNameLabel = 'main_RG30'

# CPU parameters
numberOfCpuProcessesToGenerateData = 16   # paoloPC has 16 cores
maxTasksPerChildToGenerateData = 4        # 4 is the best for paoloPC


# input state vector  parameters
numberOfActivitiesInStateVector = 6
rescaleFactorTime = 0.1
timeHorizon = 10

# random generation parameters
numberOfSimulationRunsToGenerateData =2000
numberOfSimulationRunsToTestPolicy = 1

# neural network type
neuralNetworkType = "2dimensional combined convnet"   # 1dimensional convnet, 2dimensional convnet, 1dimensional combined convnet, 2dimensional combined convnet
# for 1dimensional convnet and 2dimesnional convnet futureResourceUtilisation wont be used
if neuralNetworkType == "1dimensional convnet" or neuralNetworkType == "2dimensional convnet":
    useFutureResourceUtilisation = False

# train parameter
generateNewTrainTestValidateSets = False
importExistingNeuralNetworkModel = False
neuralNetworkModelAlreadyExists = False
numberOfEpochs = 50 #walk entire samples
learningRate = 0.1

# paths
relativePath = os.path.dirname(__file__)
absolutePathProjects = relativePath + "/database/RG30_Newdata/"

# other parameters
np.set_printoptions(precision=4)    # print precision of numpy variables

# initialise variables
numberOfActivities = None
numberOfResources = None
activitySequences = []
decisions_indexActivity = []
decisions_indexActivityPowerset = []
states = []
actions = []
# NEW:
futureResourceUtilisationMatrices = []
sumTotalDurationRandomValidateRecord = []
sumTotalDurationWithNeuralNetworkModelValidateRecord = []
sumTotalDurationWithCriticalResourceValidateRecord = []
sumTotalDurationWithShortestProcessingValidateRecord = []
sumTotalDurationWithShortestSumDurationValidateRecord = []
sumTotalDurationRandomTrainRecord = []
sumTotalDurationWithNeuralNetworkModelTrainRecord = []
sumTotalDurationWithCriticalResourceTrainRecord = []
sumTotalDurationWithShortestProcessingTrainRecord = []
sumTotalDurationWithShortestSumDurationTrainRecord = []
#sumTotalDurationRandomTestRecord = []
#sumTotalDurationWithNeuralNetworkModelTestRecord = []
#sumTotalDurationWithCriticalResourceTestRecord = []
#sumTotalDurationWithShortestProcessingTestRecord = []
#sumTotalDurationWithShortestSumDurationTestRecord = []

# read all activity sequences from database
absolutePathProjectsGlob = absolutePathProjects + "*.txt"
files = sorted(glob.glob(absolutePathProjectsGlob))

# divide all activity sequences in training and test set
numberOfFiles = len(files)

# call randomizeTrainValidateTestIndeces to generate a new train validate test split or use an already created split
indexFilesTrain, indexFilesValidate, indexFilesTest = randomizeTrainValidateTestIndeces(numberOfFiles, generateNewTrainTestValidateSets , fileNameLabel)

numberOfFilesTrain = len(indexFilesTrain)
numberOfFilesValidate = len(indexFilesValidate)
numberOfFilesTest = len(indexFilesTest)
'''
numberOfFilesTest = round(numberOfFiles * percentageOfFilesTest)
numberOfFilesTrain = numberOfFiles - numberOfFilesTest
indexFiles = list(range(0, numberOfFiles))
indexFilesTrain = []
indexFilesTest = []

# choose the first element of every set to test
for i in range(numberOfFilesTest):
    # randomIndex = random.randrange(0, len(indexFiles))
    randomIndex = i*9
    indexFilesTest.append(indexFiles[randomIndex])
    del indexFiles[randomIndex]#delete
indexFilesTrain = indexFiles
'''

# organize the read activity sequences in classes
for i in range(numberOfFiles):
    file = files[i]
    #print(file)
    # create a new activitySequence object
    currentActivitySequence = activitySequence()
    with open(file,"r") as f:
        currentActivitySequence.index = i
        currentActivitySequence.fileName = os.path.basename(f.name)
        #print(currentActivitySequence.fileName)
        # allLines = f.read()
        # print(allLines)
        next(f)
        firstLine = f.readline()    # information about numberOfActivities and numberOfResourceTypes
        firstLineDecomposed = re.split(" +", firstLine)
        numberOfActivities = (int(firstLineDecomposed[1])-2)    # the first and last dummy activity do not count
        currentActivitySequence.numberOfActivities = numberOfActivities
        # print("numberOfActivities = " + str(currentActivitySequence.numberOfActivities))
        secondLine = f.readline()   # information about total available resources
        secondLineDecomposed = re.split(" +", secondLine)
        #print(secondLineDecomposed)
        del secondLineDecomposed[0]
        secondLineDecomposed[-1]=secondLineDecomposed[-1].strip('\n')#only string have attributes of strip and delete '\n' at the end
        #print(secondLineDecomposed)
        numberOfResources = 0
        # print(len(secondLineDecomposed))
        #secondLineDecomposed=[int(secondLineDecomposed)]
        secondLineDecomposed = list(map(int,secondLineDecomposed))
        #print(secondLineDecomposed)
        for totalResources in secondLineDecomposed:
            numberOfResources += 1
            #print(numberOfResources)
            currentActivitySequence.totalResources.append(int(totalResources))
            #print(currentActivitySequence.totalResources)
        currentActivitySequence.numberOfResources = numberOfResources
        next(f)
        thirdLine = f.readline()   # information about starting activities
        thirdLineDecomposed = re.split(" +", thirdLine)
        for IdActivity in thirdLineDecomposed[7:-1]:
            currentActivitySequence.indexStartActivities.append(int(IdActivity)-2)
        #print("indexStartingActivities = " + str(currentActivitySequence.indexStartActivities))
        line = f.readline()
        while line:
            #print(line, end="")
            lineDecomposed = re.split(" +", line)
            if int(lineDecomposed[1]) == 0:
                break
            else:
                currentActivity = activity()
                currentActivity.time = int(lineDecomposed[1])
                currentActivity.requiredResources = [ int(lineDecomposed[2]),int(lineDecomposed[3]),int(lineDecomposed[4]),int(lineDecomposed[5])]
                for IdFollowingActivity in lineDecomposed[7:-1]:
                    if int(IdFollowingActivity) != numberOfActivities+2:    #if the following action is not the last dummy activity
                        currentActivity.indexFollowingActivities.append(int(IdFollowingActivity) - 2)
            currentActivitySequence.activities.append(currentActivity)
            line = f.readline()
        #add indexes to list of activities
        for j in range(len(currentActivitySequence.activities)):
            currentActivitySequence.activities[j].index = j
        #add numberOfPreviousActivities
        for Activity in currentActivitySequence.activities:
            for IndexFollowingActivity in Activity.indexFollowingActivities:
                currentActivitySequence.activities[IndexFollowingActivity].numberOfPreviousActivities += 1
    activitySequences.append(currentActivitySequence)

stateVectorLength = numberOfActivitiesInStateVector + numberOfActivitiesInStateVector * numberOfResources + numberOfResources+ timeHorizon * numberOfResources
#print(stateVectorLength)

# compute decisions: each decision corresponds to a start of an activity in the local reference system (more than one decision can be taken at once)
for i in range(0,numberOfActivitiesInStateVector):
    decisions_indexActivity.append(i)


#--------------------------------------------------------------RANDOM-----------------------------------------------------------------------------
####  GENERATE TRAINING DATA USING RANDOM DECISIONS (WITHOUT USING pool.map) ####
print('######  RANDOM DECISION ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "generateData"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    # neuralNetworkType needs to be set because states get stored accordingly in states list (vectors vs. matrices)
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences and append training data

for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesTrain[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesTrain[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesTrain[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesTrain[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesTrain[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean

    for currentStateActionPair in runSimulation_outputs[i].stateActionPairsOfBestRun:
        states.append(currentStateActionPair.state)
        actions.append(currentStateActionPair.action)
        # NEW:
        futureResourceUtilisationMatrices.append(currentStateActionPair.futureResourceUtilisationMatrix)


#correspondence best states and actions pairs --> len(states) = len(actions)
#print('state',states)
#print('actions:',actions)


####  TRAIN MODEL USING TRAINING DATA  ####
# look for existing model
#TODO: implement the application of  convnet_for_futureResourceUtilisation
print("Train neural network model")

# 1dimensional convnet without using futureResoureUtilisationMatrix
if neuralNetworkType == "1dimensional convnet":

    if importExistingNeuralNetworkModel:
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")

        else:
            neuralNetworkModel = create1dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
            # neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actionsPossibilities[0]), learningRate)
    else:
        neuralNetworkModel = create1dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
        # neuralNetworkModel = createNeuralNetworkModel(len(states[0]), len(actionsPossibilities[0]), learningRate)

    neuralNetworkModel.fit({"input": states}, {"targets": actions}, n_epoch=numberOfEpochs, snapshot_epoch=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")

# 2dimensional convnet without using futureResoureUtilisationMatrix
elif neuralNetworkType == "2dimensional convnet":
    # Turn states list into tuples
    states = np.asarray(states)
    #print("states: " + str(states))
    #print("states[0]: " + str(states[0]))
    # Reshape states
    states = states.reshape([-1, len(states[0]), len(states[0]), 1])

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            neuralNetworkModel = create2dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)
    else:
        neuralNetworkModel = create2dConvNetNeuralNetworkModel(len(states[0]), len(actions[0]), learningRate)

    neuralNetworkModel.fit({"input": states}, {"targets": actions}, n_epoch=numberOfEpochs, snapshot_epoch=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")

# combination of a 1 dimensional convnet for current state and a 2 dimensional convnet for resourceUtilisationMatrix
elif neuralNetworkType == "1dimensional combined convnet":
    placeholderVariable = 2

# combination of a 2 dimensional convnet for current state and a 2 dimensional convnet for resourceUtilisationMatrix
elif neuralNetworkType == "2dimensional combined convnet":
    # Turn states list into tuples
    states = np.asarray(states)
    # print("states: " + str(states))
    # print("states[0]: " + str(states[0]))
    # Reshape states
    states = states.reshape([-1, len(states[0]), len(states[0]), 1])

    # NEW:
    # Turn futureResourceUtilisationMatrices into tuples
    futureResourceUtilisationMatrices = np.asarray(futureResourceUtilisationMatrices)
    # Reshape futureResourceUtilisationMatrices, -1: batch_size, height(=rows):len(futureResourceUtilisationMatrices[0]), width(=columns): len(futureResourceUtilisationMatrices[0][0]), channels: 1
    futureResourceUtilisationMatrices = futureResourceUtilisationMatrices.reshape(
        [-1, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]), 1])

    if importExistingNeuralNetworkModel:
        neuralNetworkModelAlreadyExists = False
        print("check if a neural network model exists")
        if neuralNetworkModelAlreadyExists:
            print("import neural network model exists")
        else:
            # NEW:
            neuralNetworkModel = createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),len(actions[0]),learningRate,len(futureResourceUtilisationMatrices[0]),len(futureResourceUtilisationMatrices[0][0]))
    else:
        # NEW:
        neuralNetworkModel = createCombined2dConvNetNeuralNetworkModelForFutureResourceUtilisation(len(states[0]),len(actions[0]),learningRate, len(futureResourceUtilisationMatrices[0]), len(futureResourceUtilisationMatrices[0][0]))

    # NEW:
    neuralNetworkModel.fit({"input_currentState": states,
                           "input_futureResourceUtilisationMatrix": futureResourceUtilisationMatrices},
                           {"targets": actions}, n_epoch=numberOfEpochs, snapshot_epoch=500,
                           show_metric=True, run_id="trainNeuralNetworkModel")



####  CREATE BENCHMARK WITH RANDOM DECISIONS ALSO WITH VALIDATION ACTIVITY SEQUENCES  ####
print('######  RANDOM DECISION ON VALIDATE ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToGenerateData
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 1
    currentRunSimulation_input.policyType = None
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences

for i in range(numberOfFilesValidate):
    activitySequences[indexFilesValidate[i]].totalDurationMean = runSimulation_outputs[i].totalDurationMean
    activitySequences[indexFilesValidate[i]].totalDurationStandardDeviation = runSimulation_outputs[i].totalDurationStDev
    activitySequences[indexFilesValidate[i]].totalDurationMin = runSimulation_outputs[i].totalDurationMin
    activitySequences[indexFilesValidate[i]].totalDurationMax = runSimulation_outputs[i].totalDurationMax
    activitySequences[indexFilesValidate[i]].luckFactorMean = runSimulation_outputs[i].luckFactorMean
    activitySequences[indexFilesValidate[i]].trivialDecisionPercentageMean = runSimulation_outputs[i].trivialDecisionPercentageMean


#-----------------------------------------------------------------NN------------------------------------------------------------------------------
####  TEST NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON TRAIN ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesTrain[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean



####  TEST NEURAL NETWORK MODEL ON VALIDATE ACTIVITY SEQUENCES  ####
# run simulations with neural network model as decision tool (not possible to use multiprocessing -> apparently is not possible to parallelize processes on GPU)
print('###### NEURAL NETWORK MODEL ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "neuralNetworkModel"
    currentRunSimulation_input.neuralNetworkType = neuralNetworkType
    currentRunSimulation_input.decisionTool = neuralNetworkModel
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithPolicy = currentRunSimulation_output.totalDurationMean

# Xiaolei: Critical Resource, Shortest Processing Time, shortest sumDuration including successor

#---------------------------------------------------------Critical Resource----------------------------------------------------------------------------
    ####  TEST CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource = runSimulation_outputs[i].totalDurationMean


####  TEST CRITICAL RESOURCE METHOD ON VALIDATE ACTIVITY SEQUENCES  ####
print('###### CRITICAL RESOURCE METHOD ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "most critical resource"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithCriticalResource = currentRunSimulation_output.totalDurationMean

# ---------------------------------------------------------Shortest Processing Time----------------------------------------------------------------------------
####  TEST SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST PROCESSING TIME METHOD ON VALIDATE ACTIVITY SEQUENCES  ####
print('###### SHORTEST PROCESSING TIME METHOD ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest processing time"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithShortestProcessingTime = currentRunSimulation_output.totalDurationMean

 # ---------------------------------------------------------shortest sumDuration including successor----------------------------------------------------------------------------
####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON TRAIN ACTIVITY SEQUENCES  ######')
runSimulation_inputs = []
for i in range(numberOfFilesTrain):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesTrain[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    runSimulation_inputs.append(currentRunSimulation_input)

pool = mp.Pool(processes=numberOfCpuProcessesToGenerateData)

runSimulation_outputs = pool.map(runSimulation, runSimulation_inputs)
# assign simulation results to activity sequences
for i in range(numberOfFilesTrain):
    activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration = runSimulation_outputs[i].totalDurationMean

####  TEST SHORTEST SUMDURATION INCLUDING SUCCESSOR TIME METHOD ON VALIDATE ACTIVITY SEQUENCES  ####
print('###### SHORTEST SUMDURATION INCLUDING SUCCESSOR METHOD ON VALIDATE ACTIVITY SEQUENCES  ######')
for i in range(numberOfFilesValidate):
    currentRunSimulation_input = runSimulation_input()
    currentRunSimulation_input.activitySequence = activitySequences[indexFilesValidate[i]]
    currentRunSimulation_input.numberOfSimulationRuns = numberOfSimulationRunsToTestPolicy
    currentRunSimulation_input.timeDistribution = timeDistribution
    currentRunSimulation_input.purpose = "testPolicy"
    currentRunSimulation_input.randomDecisionProbability = 0
    currentRunSimulation_input.policyType = "shortest sumDuration including successor"
    currentRunSimulation_input.neuralNetworkType = None
    currentRunSimulation_input.decisionTool = None
    currentRunSimulation_input.numberOfResources = numberOfResources
    currentRunSimulation_input.numberOfActivitiesInStateVector = numberOfActivitiesInStateVector
    currentRunSimulation_input.stateVectorLength = stateVectorLength
    currentRunSimulation_input.decisions_indexActivity = decisions_indexActivity
    currentRunSimulation_input.rescaleFactorTime = rescaleFactorTime
    currentRunSimulation_input.numberOfActivities = numberOfActivities
    currentRunSimulation_input.timeHorizon = timeHorizon

    currentRunSimulation_output = runSimulation(currentRunSimulation_input)

    activitySequences[indexFilesValidate[i]].totalDurationWithShortestSumDuration = currentRunSimulation_output.totalDurationMean

#------------------------------------------------------EVALUATION-----------------------------------------------------------------------------
####  EVALUATION OF RESULTS OF TRAIN ACTIVITY SEQUENCES  ####
sumTotalDurationRandomTrain = 0
sumTotalDurationWithCriticalResourceTrain = 0
sumTotalDurationWithShortestProcessingTrain = 0
sumTotalDurationWithNeuralNetworkModelTrain = 0
sumTotalDurationWithShortestSumDurationTrain = 0

for i in range(numberOfFilesTrain):
    sumTotalDurationRandomTrain += activitySequences[indexFilesTrain[i]].totalDurationMean
    sumTotalDurationRandomTrain = round(sumTotalDurationRandomTrain,4)
    sumTotalDurationWithNeuralNetworkModelTrain += activitySequences[indexFilesTrain[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceTrain += activitySequences[indexFilesTrain[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationTrain += activitySequences[indexFilesTrain[i]].totalDurationWithShortestSumDuration

sumTotalDurationRandomTrainRecord.append(sumTotalDurationRandomTrain)
sumTotalDurationWithNeuralNetworkModelTrainRecord.append(sumTotalDurationWithNeuralNetworkModelTrain)
sumTotalDurationWithCriticalResourceTrainRecord.append(sumTotalDurationWithCriticalResourceTrain)
sumTotalDurationWithShortestProcessingTrainRecord.append(sumTotalDurationWithShortestProcessingTrain)
sumTotalDurationWithShortestSumDurationTrainRecord.append(sumTotalDurationWithShortestSumDurationTrain)

####  EVALUATION OF NN RESULTS OF VALIDATE ACTIVITY SEQUENCES  ####
sumTotalDurationRandomValidate = 0
sumTotalDurationWithNeuralNetworkModelValidate = 0
sumTotalDurationWithCriticalResourceValidate = 0
sumTotalDurationWithShortestProcessingValidate = 0
sumTotalDurationWithShortestSumDurationValidate = 0

for i in range(numberOfFilesValidate):
    sumTotalDurationRandomValidate += activitySequences[indexFilesValidate[i]].totalDurationMean
    sumTotalDurationRandomValidate = round(sumTotalDurationRandomValidate,4)
    sumTotalDurationWithNeuralNetworkModelValidate += activitySequences[indexFilesValidate[i]].totalDurationWithPolicy
    sumTotalDurationWithCriticalResourceValidate += activitySequences[indexFilesValidate[i]].totalDurationWithCriticalResource
    sumTotalDurationWithShortestProcessingValidate += activitySequences[indexFilesValidate[i]].totalDurationWithShortestProcessingTime
    sumTotalDurationWithShortestSumDurationValidate += activitySequences[indexFilesValidate[i]].totalDurationWithShortestSumDuration


sumTotalDurationRandomValidateRecord.append(sumTotalDurationRandomValidate)
sumTotalDurationWithNeuralNetworkModelValidateRecord.append(sumTotalDurationWithNeuralNetworkModelValidate)
sumTotalDurationWithCriticalResourceValidateRecord.append(sumTotalDurationWithCriticalResourceValidate)
sumTotalDurationWithShortestProcessingValidateRecord.append(sumTotalDurationWithShortestProcessingValidate)
sumTotalDurationWithShortestSumDurationValidateRecord.append(sumTotalDurationWithShortestSumDurationValidate)


print("neuralNetworkType: " + neuralNetworkType)
print("sumTotalDurationRandomTrain = " + str(sumTotalDurationRandomTrain))
print("sumTotalDurationWithNeuralNetworkModelTrain = " + str(sumTotalDurationWithNeuralNetworkModelTrain))
print("sumTotalDurationWithCriticalResourceTrain = " + str(sumTotalDurationWithCriticalResourceTrain))
print("sumTotalDurationWithShortestProcessingTrain = " + str(sumTotalDurationWithShortestProcessingTrain))
print("sumTotalDurationWithShortestSumDurationTrain = " + str(sumTotalDurationWithShortestSumDurationTrain))
print("sumTotalDurationRandomValidate = " + str(sumTotalDurationRandomValidate))
print("sumTotalDurationWithNeuralNetworkModelValidate = " + str(sumTotalDurationWithNeuralNetworkModelValidate))
print("sumTotalDurationWithCriticalResourceValidate = " + str(sumTotalDurationWithCriticalResourceValidate))
print("sumTotalDurationWithShortestProcessingValidate = " + str(sumTotalDurationWithShortestProcessingValidate))
print("sumTotalDurationWithShortestSumDurationValidate = " + str(sumTotalDurationWithShortestSumDurationValidate))


# compute computation time
t_end = time.time()
t_computation = t_end - t_start
print("t_computation = " + str(t_computation))


#write ouput to excel
wb = Workbook()
ws = wb.create_sheet('Durations_psplib',0)

alignCenter = Alignment(horizontal='center')

ws['A1'] = 'Durations'
ws['B1'] = 'Computation time'
ws['C1'].value = t_computation
ws.merge_cells('A2:E2')
ws.merge_cells('F2:J2')
ws['A2'] = 'durations on train set'
ws['A2'].alignment = alignCenter
ws['F2'] = 'durations on validation set'
ws['F2'].alignment = alignCenter

ws['A3'] = 'Random'
ws['B3'] = 'NeuralNetworkModel'
ws['C3'] = 'CriticalResource'
ws['D3'] = 'ShortestProcessing'
ws['E3'] = 'ShortestSumDuration'
ws['F3'] = 'Random'
ws['G3'] = 'NeuralNetworkModel'
ws['H3'] = 'CriticalResource'
ws['I3'] = 'ShortestProcessing'
ws['J3'] = 'ShortestSumDuration'

ws['A4'].value = sumTotalDurationRandomTrain
ws['B4'].value = sumTotalDurationWithNeuralNetworkModelTrain
ws['C4'].value = sumTotalDurationWithCriticalResourceTrain
ws['D4'].value = sumTotalDurationWithShortestProcessingTrain
ws['E4'].value = sumTotalDurationWithShortestSumDurationTrain
ws['F4'].value = sumTotalDurationRandomValidate
ws['G4'].value = sumTotalDurationWithNeuralNetworkModelValidate
ws['H4'].value = sumTotalDurationWithCriticalResourceValidate
ws['I4'].value = sumTotalDurationWithShortestProcessingValidate
ws['J4'].value = sumTotalDurationWithShortestSumDurationValidate




ws.column_dimensions['A'].width = 10.0
ws.column_dimensions['B'].width = 18.0
ws.column_dimensions['C'].width = 14.0
ws.column_dimensions['D'].width = 18.0
ws.column_dimensions['E'].width = 19.0
ws.column_dimensions['F'].width = 10.0
ws.column_dimensions['G'].width = 18.0
ws.column_dimensions['H'].width = 14.0
ws.column_dimensions['I'].width = 18.0
ws.column_dimensions['J'].width = 19.0

ws['A3'].alignment = alignCenter
ws['B3'].alignment = alignCenter
ws['C3'].alignment = alignCenter
ws['D3'].alignment = alignCenter
ws['E3'].alignment = alignCenter
ws['F3'].alignment = alignCenter
ws['G3'].alignment = alignCenter
ws['H3'].alignment = alignCenter
ws['I3'].alignment = alignCenter
ws['J3'].alignment = alignCenter

ws['A4'].alignment = alignCenter
ws['B4'].alignment = alignCenter
ws['C4'].alignment = alignCenter
ws['D4'].alignment = alignCenter
ws['E4'].alignment = alignCenter
ws['F4'].alignment = alignCenter
ws['G4'].alignment = alignCenter
ws['H4'].alignment = alignCenter
ws['I4'].alignment = alignCenter
ws['J4'].alignment = alignCenter


wb.save(relativePath + "/durations_rg30.xlsx")
